"""Parse the agency's "Enterprise Reporting [Compiled]" workbook.

Marketing Services compiles validated (human-verified) marketing performance
for Enterprise practices each month. Those numbers supersede Omni's
attribution for the MBR marketing section, so reports for these practices
show the same figures the agency reported.

Summary-sheet layout (one row per practice-offer):
  - Column A: "<medspa_id> <practice name>"  (e.g. "1734 Epitome of Youth")
  - "Offer" column: campaign name; a synthetic "Totals" row precedes the
    per-offer rows for practices running multiple offers (we sum offers
    ourselves and ignore Totals rows)
  - "Data" column: "Validated" rows carry the numbers; "Omni" rows are
    empty placeholders
"""

import re
from pathlib import Path


# Sheet header → internal field. Headers are matched case-insensitively by
# prefix so small wording tweaks ("Revenue" vs "Revenu") don't break parsing.
_COLUMNS = {
    "medspa name": "medspa",
    "offer": "offer",
    "data": "data_source",
    "spend": "ad_spend",
    "meta leads": "leads",
    "new patient leads": "new_patient_leads",
    "new patient appts booked": "booked",
    "new patient appts completed": "completed",
    "new patient first visit revenue": "revenue",
    "total revenue generated": "total_revenue_all_clients",
    "first visit new patient roi": "first_visit_roi",
    "lead to booked appt conversion": "lead_to_booking_rate",
    "notes": "notes",
}

_ID_NAME_RE = re.compile(r"^\s*(\d+)\s+(.+?)\s*$", re.DOTALL)


def _num(v, cast=float):
    if v is None or v == "":
        return None
    try:
        return cast(float(v))
    except (TypeError, ValueError):
        return None


def parse_enterprise_workbook(path: str) -> dict:
    """Parse the compiled workbook's Summary sheet.

    Returns {"practices": {medspa_id_str: record}, "skipped": [names]}.
    A record aggregates the practice's Validated offer rows:
      practice_name, ad_spend, leads, booked, completed, revenue,
      total_revenue_all_clients, first_visit_roi, lead_to_booking_rate,
      first_visit_aov, campaigns: [{campaign_name, ad_spend, leads,
      booked, completed, revenue}]
    Practices whose rows carry no numbers at all are listed in "skipped"
    (their reports keep Omni's marketing data).
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    sheet_name = next((n for n in wb.sheetnames if n.strip().lower() == "summary"),
                      wb.sheetnames[0])
    ws = wb[sheet_name]

    # Locate the header row (the one containing "MedSpa Name")
    header_row = None
    col_map = {}
    for row in ws.iter_rows(min_row=1, max_row=10):
        cells = [(c.column, str(c.value).strip().lower())
                 for c in row if c.value is not None]
        if any(v == "medspa name" for _, v in cells):
            header_row = row[0].row
            for col_idx, text in cells:
                for prefix, field in _COLUMNS.items():
                    if text.startswith(prefix) and field not in col_map:
                        col_map[field] = col_idx
            break
    if header_row is None or "medspa" not in col_map:
        raise ValueError(
            f"Could not find a 'MedSpa Name' header row in sheet "
            f"{sheet_name!r} — is this the Enterprise Reporting workbook?")

    def cell(row_vals, field):
        idx = col_map.get(field)
        return row_vals[idx - 1] if idx and idx - 1 < len(row_vals) else None

    practices: dict = {}
    skipped: list = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        raw_name = cell(row, "medspa")
        offer = cell(row, "offer")
        source = str(cell(row, "data_source") or "").strip().lower()

        if not raw_name or not str(raw_name).strip():
            continue
        m = _ID_NAME_RE.match(str(raw_name))
        if not m:
            skipped.append(str(raw_name).strip())
            continue
        # Synthetic aggregate rows — we sum the offer rows ourselves
        if str(offer or "").strip().lower() == "totals":
            continue
        # "Omni" placeholder rows carry no validated numbers
        if source == "omni":
            continue

        spend = _num(cell(row, "ad_spend"))
        leads = _num(cell(row, "leads"), int)
        booked = _num(cell(row, "booked"), int)
        completed = _num(cell(row, "completed"), int)
        revenue = _num(cell(row, "revenue"))
        total_rev = _num(cell(row, "total_revenue_all_clients"))
        if all(v is None for v in (spend, leads, booked, completed,
                                   revenue, total_rev)):
            skipped.append(str(raw_name).strip().replace("\n", " "))
            continue

        medspa_id = str(int(m.group(1)))
        name = " ".join(m.group(2).split())
        rec = practices.setdefault(medspa_id, {
            "practice_name": name,
            "ad_spend": 0.0, "leads": 0, "new_patient_leads": 0,
            "booked": 0, "completed": 0,
            "revenue": 0.0, "total_revenue_all_clients": 0.0,
            "campaigns": [],
        })
        rec["ad_spend"] += spend or 0.0
        rec["leads"] += leads or 0
        rec["new_patient_leads"] += _num(cell(row, "new_patient_leads"), int) or 0
        rec["booked"] += booked or 0
        rec["completed"] += completed or 0
        rec["revenue"] += revenue or 0.0
        rec["total_revenue_all_clients"] += total_rev or 0.0
        rec["campaigns"].append({
            "campaign_name": " ".join(str(offer or "Campaign").split()),
            "ad_spend": spend or 0.0,
            "leads": leads or 0,
            "booked": booked or 0,
            "completed": completed or 0,
            "revenue": revenue or 0.0,
        })

    # Derived ratios — computed from the aggregated sums so multi-offer
    # practices match the workbook's own "Totals" rows.
    for rec in practices.values():
        rec["ad_spend"] = round(rec["ad_spend"], 2)
        rec["revenue"] = round(rec["revenue"], 2)
        rec["total_revenue_all_clients"] = round(rec["total_revenue_all_clients"], 2)
        rec["first_visit_roi"] = (round(rec["revenue"] / rec["ad_spend"], 2)
                                  if rec["ad_spend"] > 0 else None)
        # The workbook's conversion rate is booked / NEW-PATIENT leads
        npl = rec.pop("new_patient_leads", 0)
        rec["lead_to_booking_rate"] = (rec["booked"] / npl if npl > 0 else None)
        rec["first_visit_aov"] = (round(rec["revenue"] / rec["completed"], 2)
                                  if rec["completed"] > 0 else None)

    return {"practices": practices,
            "skipped": sorted(set(skipped))}


def reconcile_with_omni(parsed: dict, api_key: str) -> dict:
    """Verify parsed medspa ids against Omni and fix mis-parsed ones.

    The workbook's leading number is *usually* the medspa id, but some
    practices have a number in their business name (e.g. the medspa named
    "424 Cosmetic Dermatology" is id 1790). For every parsed id that
    doesn't exist in Omni, retry the lookup by name; single-hit matches are
    re-keyed, everything else is dropped and reported as unmatched so a
    human can fix the sheet.

    Mutates `parsed` in place. Returns
    {"verified": bool, "remapped": {sheet_row: new_id}, "unmatched": [rows]}.
    """
    import copy
    from .omni_loader import (_resolve_mbr_dashboard, _run_query,
                              _ensure_filters)

    ID_F = "dbt__moxie_medspas_mart.medspa_id"
    NAME_F = "dbt__moxie_medspas_mart.medspa_name"

    dash = _resolve_mbr_dashboard(api_key)
    chassis = next((q["query"] for q in dash.get("queries", [])
                    if q.get("name") == "Medspa Name" and q.get("query")), None)
    if chassis is None:
        return {"verified": False, "remapped": {}, "unmatched": []}

    def _prep(q):
        if not isinstance(q.get("fields"), list):
            q["fields"] = []
        for f in (ID_F, NAME_F):
            if f not in q["fields"]:
                q["fields"].append(f)
        q["limit"] = 100
        return q

    def by_ids(ids):
        q = _prep(copy.deepcopy(chassis))
        _ensure_filters(q)[ID_F] = {
            "kind": "EQUALS", "type": "number", "is_inclusive": False,
            "values": ids, "is_negative": False}
        r = _run_query(q, api_key)
        return {int(m): n for m, n in zip(r.get(ID_F, []), r.get(NAME_F, []))
                if m is not None}

    def by_name(name):
        q = _prep(copy.deepcopy(chassis))
        _ensure_filters(q)[NAME_F] = {
            "kind": "CONTAINS", "type": "string",
            "values": [name], "is_negative": False}
        r = _run_query(q, api_key)
        return [(int(m), n) for m, n in zip(r.get(ID_F, []), r.get(NAME_F, []))
                if m is not None]

    practices = parsed["practices"]
    found = by_ids([int(i) for i in practices])
    remapped, unmatched = {}, []
    for mid in list(practices):
        rec = practices[mid]
        if int(mid) in found:
            rec["omni_name"] = found[int(mid)]
            continue
        sheet_row = f"{mid} {rec['practice_name']}"
        hits = by_name(sheet_row) or by_name(rec["practice_name"])
        if len(hits) == 1:
            new_id, omni_name = hits[0]
            rec["omni_name"] = omni_name
            practices[str(new_id)] = practices.pop(mid)
            remapped[sheet_row] = str(new_id)
        else:
            unmatched.append(sheet_row)
            practices.pop(mid)
    return {"verified": True, "remapped": remapped, "unmatched": unmatched}
